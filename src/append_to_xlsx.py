import os
import pandas as pd
import numpy as np
from openpyxl.styles import NamedStyle


def append_df_to_excel(
    filename,
    df,
    sheet_name=None,
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into a new sheet. If the file does not exist, then it will be created.

    Parameters:
    filename : File path or existing ExcelWriter (Example: '/path/to/file.xlsx')
    df : DataFrame to save to workbook
    sheet_name : Name of the sheet which will contain DataFrame. If None, then use default name ('Sheet1', 'Sheet2', etc.)
    startrow : Upper left cell row to dump data frame. Per default (startrow=None) calculate the last row in the existing DF and write to the next row...
    truncate_sheet : Truncate (remove and recreate) [sheet_name] before writing DataFrame to Excel file
    to_excel_kwargs : Arguments which will be passed to `DataFrame.to_excel()` [can be a dictionary]
    """

    # Ensure a default is set for index
    to_excel_kwargs["index"] = to_excel_kwargs.get("index", False)

    # If file does not exist, then create it
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs
        )
    else:
        # If file exists, append to it
        with pd.ExcelWriter(filename, engine="openpyxl", mode="a") as writer:
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # Find worksheet by name and remove it
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])

            # Write the DataFrame
            df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

            # Apply date format to date columns
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")

            for col in df.select_dtypes(include=[np.datetime64, "datetime"]).columns:
                col_idx = df.columns.get_loc(col) + 1  # +1 because Excel is 1-indexed
                for row in range(2, len(df) + 2):  # +2 because of header and 1-indexed
                    worksheet.cell(column=col_idx, row=row).style = date_style

            # Save the workbook
            workbook.save(filename)
