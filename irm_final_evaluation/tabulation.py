""""Extract Valuce Counts of Survey Data"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# Load your dataframe here
dataset = pd.read_excel("irm_survey.xlsx")

# Create a new Workbook
wb = Workbook()
ws = wb.active
ws.title = "Value Counts"

# Iterate over each column in the DataFrame
for col in dataset.columns:
    # Perform value counts including NaN values
    value_counts = dataset[col].value_counts(dropna=False)

    # Create a DataFrame with counts and relative frequencies
    value_counts_df = pd.DataFrame({
        '#': range(1, len(value_counts) + 1),
        'Value': value_counts.index,
        'Count': value_counts.values,
        'Relative Frequency': (value_counts / value_counts.sum()).round(2) 
    })

    # Replace NaN with 'Missing values' in the 'Value' column
    value_counts_df['Value'] = value_counts_df['Value'].fillna('Missing values')

    # Enhance the column name
    col_name = col.replace('_', ' ').upper()

    # Write column name
    ws.append([col_name])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)  # Make column name bold

    # Write beader row for value, count, and relative frequency
    ws.append(["#" ,"Value", "Count", "Relative Frequency"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)

    # Write each value, its count, and relative frequency
    for _, row in value_counts_df.iterrows():
        ws.append([row['#'] ,row['Value'], row['Count'], row['Relative Frequency']])
    
    # Add a row for the sum of counts and relative requencies
    ws.append(["", "Sum", value_counts_df['Count'].sum(), value_counts_df['Relative Frequency'].sum()])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)

    # Add an empty row after each column value counts to improve readability
    ws.append([])

# Save the workbook to an Excel file
wb.save("value_counts_summary.xlsx")

print("Value counts saved to value_counts-summary.xlsx")
